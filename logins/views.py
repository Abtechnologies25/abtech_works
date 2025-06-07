from django.shortcuts import render, redirect,get_object_or_404
from datetime import datetime
import pandas as pd
from django.http import HttpResponse
from django.contrib.auth import login, authenticate, logout
from django.views.decorators.cache import cache_control
from .forms import UserRegisterForm, UserLoginForm
from django.contrib.auth.decorators import login_required
from .models import *
from .forms import *
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from django.contrib import messages
import xlwt
from itertools import chain
from collections import defaultdict
def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  # ❌ Prevent login until admin approval
            user.save()
            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request, 'logins/register.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserLoginForm()
    return render(request, 'logins/login.html', {'form': form})

def admin_dashboard(request):
    # purchase_orders = PurchaseOrder.objects.all()
    # dealers = Dealer.objects.all()
    user_registrations = User.objects.all()
    BRANCHES = ['Nagercoil', 'Tirunelveli', 'Pudukottai', 'Chennai']
    nagercoil_projects = NagercoilProjectRegistration.objects.all()
    tirunelveli_projects = TirunelveliProjectRegistration.objects.all()
    pudukottai_projects = PudukottaiProjectRegistration.objects.all()
    chennai_projects = ChennaiProjectRegistration.objects.all()
    nagercoil_internships = NagercoilInternshipRegistration.objects.all()
    tirunelveli_internships = TirunelveliInternshipRegistration.objects.all()
    pudukottai_internships = PudukottaiInternshipRegistration.objects.all()
    chennai_internships = ChennaiInternshipRegistration.objects.all()
    nagercoil_publications = NagercoilPublicationRegistration.objects.all()
    tirunelveli_publications = TirunelveliPublicationRegistration.objects.all()
    pudukottai_publications = PudukottaiPublicationRegistration.objects.all()
    chennai_publications = ChennaiPublicationRegistration.objects.all()
    nagercoil_income_expenses = NagercoilDailyIncomeExpenditure.objects.all()
    tirunelveli_income_expenses = TirunelveliDailyIncomeExpenditure.objects.all()
    pudukottai_income_expenses = PudukottaiDailyIncomeExpenditure.objects.all()
    chennai_income_expenses = ChennaiDailyIncomeExpenditure.objects.all()
    nagercoil_payment_vouchers = NagercoilPaymentVoucher.objects.all()
    tirunelveli_payment_vouchers = TirunelveliPaymentVoucher.objects.all()
    pudukottai_payment_vouchers = PudukottaiPaymentVoucher.objects.all()
    chennai_payment_vouchers = ChennaiPaymentVoucher.objects.all()
    phd_bills = {
        'Nagercoil': NagercoilPhdBill.objects.all(),
        'Tirunelveli': TirunelveliPhdBill.objects.all(),
        'Pudukottai': PudukottaiPhdBill.objects.all(),
        'Chennai': ChennaiPhdBill.objects.all(),
    }

    internship_bills = {
        'Nagercoil': NagercoilInternshipBill.objects.all(),
        'Tirunelveli': TirunelveliInternshipBill.objects.all(),
        'Pudukottai': PudukottaiInternshipBill.objects.all(),
        'Chennai': ChennaiInternshipBill.objects.all(),
    }

    project_bills = {
        'Nagercoil': NagercoilProjectBill.objects.all(),
        'Tirunelveli': TirunelveliProjectBill.objects.all(),
        'Pudukottai': PudukottaiProjectBill.objects.all(),
        'Chennai': ChennaiProjectBill.objects.all(),
    }

    journal_bills = {
        'Nagercoil': NagercoilJournalBill.objects.all(),
        'Tirunelveli': TirunelveliJournalBill.objects.all(),
        'Pudukottai': PudukottaiJournalBill.objects.all(),
        'Chennai': ChennaiJournalBill.objects.all(),
    }

    sharing_bills = {
        'Nagercoil': NagercoilSharingBill.objects.all(),
        'Tirunelveli': TirunelveliSharingBill.objects.all(),
        'Pudukottai': PudukottaiSharingBill.objects.all(),
        'Chennai': ChennaiSharingBill.objects.all(),
    }

    patent_bills = {
        'Nagercoil': NagercoilPatentBill.objects.all(),
        'Tirunelveli': TirunelveliPatentBill.objects.all(),
        'Pudukottai': PudukottaiPatentBill.objects.all(),
        'Chennai': ChennaiPatentBill.objects.all(),
    }
    work_status_records = WorkStatus.objects.all()
    DEPARTMENT_CHOICES = [
    'journal team',
    'Research work development team',
    'software project development team',
    'hardware project development team',
    'Technical team',
    'Admin team',
]
    context = {
        'user_registrations': user_registrations,
        'nagercoil_phds': NagercoilPhdRegistration.objects.all(),
        'tirunelveli_phds': TirunelveliPhdRegistration.objects.all(),
        'pudukottai_phds': PudukottaiPhdRegistration.objects.all(),
        'chennai_phds': ChennaiPhdRegistration.objects.all(),
        'nagercoil_projects': nagercoil_projects,
        'tirunelveli_projects': tirunelveli_projects,
        'pudukottai_projects': pudukottai_projects,
        'chennai_projects': chennai_projects,
        'nagercoil_internships': nagercoil_internships,
        'tirunelveli_internships': tirunelveli_internships,
        'pudukottai_internships': pudukottai_internships,
        'chennai_internships': chennai_internships,
        'nagercoil_publications': nagercoil_publications,
        'tirunelveli_publications': tirunelveli_publications,
        'pudukottai_publications': pudukottai_publications,
        'chennai_publications': chennai_publications,
        'nagercoil_income_expenses': nagercoil_income_expenses,
        'tirunelveli_income_expenses': tirunelveli_income_expenses,
        'pudukottai_income_expenses': pudukottai_income_expenses,
        'chennai_income_expenses': chennai_income_expenses,
        'nagercoil_payment_vouchers': nagercoil_payment_vouchers,
        'tirunelveli_payment_vouchers': tirunelveli_payment_vouchers,
        'pudukottai_payment_vouchers': pudukottai_payment_vouchers,
        'chennai_payment_vouchers': chennai_payment_vouchers,
        'phd_bills': phd_bills,
        'internship_bills': internship_bills,
        'project_bills': project_bills,
        'journal_bills': journal_bills,
        'sharing_bills': sharing_bills,
        'patent_bills': patent_bills,
        # 'work_status_records': work_status_records,
    'nagercoil_work_status': work_status_records.filter(branch='Nagercoil').order_by('-date'),
    'tirunelveli_work_status': work_status_records.filter(branch='Tirunelveli').order_by('-date'),
    'pudukottai_work_status': work_status_records.filter(branch='Pudukottai').order_by('-date'),
    'chennai_work_status': work_status_records.filter(branch='Chennai').order_by('-date'),
    'departments': DEPARTMENT_CHOICES,
    'branches': BRANCHES,
    # 'purchase_orders': purchase_orders,
    # 'dealers': dealers,
    }

    return render(request, 'logins/admin_dashboard.html', context)

def admin_purchase_orders(request):
    orders = PurchaseOrder.objects.all()
    return render(request, 'logins/admin_purchase_orders.html', {'orders': orders})

def admin_dealer_list(request):
    dealers = Dealer.objects.all()
    return render(request, 'logins/admin_dealer_list.html', {'dealers': dealers})

def admin_dealer_orders(request, dealer_id):
    dealer = Dealer.objects.get(id=dealer_id)
    orders = dealer.dealerpurchaseorder_set.all()
    return render(request, 'logins/admin_dealer_orders.html', {'dealer': dealer, 'orders': orders})

def nagercoil_admin_dashboard(request):
    nagercoil_projects = NagercoilProjectRegistration.objects.all()
    nagercoil_internships = NagercoilInternshipRegistration.objects.all()
    nagercoil_publications = NagercoilPublicationRegistration.objects.all()
    nagercoil_income_expenses = NagercoilDailyIncomeExpenditure.objects.all()
    nagercoil_payment_vouchers = NagercoilPaymentVoucher.objects.all()
    context = {
        'nagercoil_phds': NagercoilPhdRegistration.objects.all(),
        'nagercoil_projects': nagercoil_projects,
        'nagercoil_internships': nagercoil_internships,
        'nagercoil_publications': nagercoil_publications,
        'nagercoil_income_expenses': nagercoil_income_expenses,
        'nagercoil_payment_vouchers': nagercoil_payment_vouchers,
    }
    return render(request, 'logins/nagercoil_admin_dashboard.html', context)
def chennai_admin_dashboard(request):
    chennai_projects = ChennaiProjectRegistration.objects.all()
    chennai_internships = ChennaiInternshipRegistration.objects.all()
    chennai_publications = ChennaiPublicationRegistration.objects.all()
    chennai_income_expenses = ChennaiDailyIncomeExpenditure.objects.all()
    chennai_payment_vouchers = ChennaiPaymentVoucher.objects.all()
    context = {
        'chennai_phds': ChennaiPhdRegistration.objects.all(),
        'chennai_projects': chennai_projects,
        'chennai_internships': chennai_internships,
        'chennai_publications': chennai_publications,
        'chennai_income_expenses': chennai_income_expenses,
        'chennai_payment_vouchers': chennai_payment_vouchers,
    }
    return render(request, 'logins/chennai_admin_dashboard.html', context)

def tirunelveli_admin_dashboard(request):
    tirunelveli_projects = TirunelveliProjectRegistration.objects.all()
    tirunelveli_internships = TirunelveliInternshipRegistration.objects.all()
    tirunelveli_publications = TirunelveliPublicationRegistration.objects.all()
    tirunelveli_income_expenses = TirunelveliDailyIncomeExpenditure.objects.all()
    tirunelveli_payment_vouchers = TirunelveliPaymentVoucher.objects.all()
    context = {
    'tirunelveli_phds': TirunelveliPhdRegistration.objects.all(),
    'tirunelveli_projects': tirunelveli_projects,
    'tirunelveli_internships': tirunelveli_internships,
    'tirunelveli_publications': tirunelveli_publications,
    'tirunelveli_income_expenses': tirunelveli_income_expenses,
    'tirunelveli_payment_vouchers': tirunelveli_payment_vouchers,
    }
    return render(request, 'logins/tirunelveli_admin_dashboard.html', context)

def pudukottai_admin_dashboard(request):
    pudukottai_projects = PudukottaiProjectRegistration.objects.all() 
    pudukottai_internships = PudukottaiInternshipRegistration.objects.all()
    pudukottai_publications = PudukottaiPublicationRegistration.objects.all()
    pudukottai_income_expenses = PudukottaiDailyIncomeExpenditure.objects.all()
    pudukottai_payment_vouchers = PudukottaiPaymentVoucher.objects.all()
    context = {
        'pudukottai_phds': PudukottaiPhdRegistration.objects.all(),
        'pudukottai_projects': pudukottai_projects,
        'pudukottai_internships': pudukottai_internships,
        'pudukottai_publications': pudukottai_publications,
        'pudukottai_income_expenses': pudukottai_income_expenses,
        'pudukottai_payment_vouchers': pudukottai_payment_vouchers,
    }
    return render(request, 'logins/pudukottai_admin_dashboard.html', context)
def billwise_admin_dashboard(request):
    BRANCHES = ['Nagercoil', 'Tirunelveli', 'Pudukottai', 'Chennai']
    phd_bills = {
        'Nagercoil': NagercoilPhdBill.objects.all(),
        'Tirunelveli': TirunelveliPhdBill.objects.all(),
        'Pudukottai': PudukottaiPhdBill.objects.all(),
        'Chennai': ChennaiPhdBill.objects.all(),
    }

    internship_bills = {
        'Nagercoil': NagercoilInternshipBill.objects.all(),
        'Tirunelveli': TirunelveliInternshipBill.objects.all(),
        'Pudukottai': PudukottaiInternshipBill.objects.all(),
        'Chennai': ChennaiInternshipBill.objects.all(),
    }

    project_bills = {
        'Nagercoil': NagercoilProjectBill.objects.all(),
        'Tirunelveli': TirunelveliProjectBill.objects.all(),
        'Pudukottai': PudukottaiProjectBill.objects.all(),
        'Chennai': ChennaiProjectBill.objects.all(),
    }

    journal_bills = {
        'Nagercoil': NagercoilJournalBill.objects.all(),
        'Tirunelveli': TirunelveliJournalBill.objects.all(),
        'Pudukottai': PudukottaiJournalBill.objects.all(),
        'Chennai': ChennaiJournalBill.objects.all(),
    }

    sharing_bills = {
        'Nagercoil': NagercoilSharingBill.objects.all(),
        'Tirunelveli': TirunelveliSharingBill.objects.all(),
        'Pudukottai': PudukottaiSharingBill.objects.all(),
        'Chennai': ChennaiSharingBill.objects.all(),
    }

    patent_bills = {
        'Nagercoil': NagercoilPatentBill.objects.all(),
        'Tirunelveli': TirunelveliPatentBill.objects.all(),
        'Pudukottai': PudukottaiPatentBill.objects.all(),
        'Chennai': ChennaiPatentBill.objects.all(),
    }
    context={
        'phd_bills': phd_bills,
        'internship_bills': internship_bills,
        'project_bills': project_bills,
        'journal_bills': journal_bills,
        'sharing_bills': sharing_bills,
        'patent_bills': patent_bills,
    }
    return render(request, 'logins/billwise_admin_dashboard.html', context)
def workstatus_admin_dashboard(request):
    BRANCHES = ['Nagercoil', 'Tirunelveli', 'Pudukottai', 'Chennai']
    DEPARTMENT_CHOICES = [
    'journal team',
    'Research work development team',
    'software project development team',
    'hardware project development team',
    'Technical team',
    'Admin team',
    ]
    work_status_records = WorkStatus.objects.all()

    context={
    'nagercoil_work_status': work_status_records.filter(branch='Nagercoil').order_by('-date'),
    'tirunelveli_work_status': work_status_records.filter(branch='Tirunelveli').order_by('-date'),
    'pudukottai_work_status': work_status_records.filter(branch='Pudukottai').order_by('-date'),
    'chennai_work_status': work_status_records.filter(branch='Chennai').order_by('-date'),
    'departments': DEPARTMENT_CHOICES,
    'branches': BRANCHES,
    }
    return render(request, 'logins/workstatus_admin_dashboard.html', context)
@login_required
def approve_user(request, user_id):
    if not request.user.is_superuser:
        return redirect('login')

    user = get_object_or_404(User, id=user_id)
    user.is_approved = True  # ✅ Mark as approved
    user.is_active = True  # ✅ Allow login
    user.save()

    return redirect('admin_dashboard')

@login_required
def reject_user(request, user_id):
    if not request.user.is_superuser:
        return redirect('login')

    user = get_object_or_404(User, id=user_id)
    user.delete()  # ❌ Remove rejected users
    return redirect('admin_dashboard')

@login_required
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.error(request, f"User {user.username} has been deleted.")
    return redirect('admin_dashboard')

def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        user_form = UserEditForm(request.POST, instance=user)
        password_form = UserPasswordEditForm(user, request.POST)

        if user_form.is_valid() and password_form.is_valid():
            user_form.save()
            password_form.save()
            return redirect('admin_dashboard')  # Redirect to admin dashboard after editing
    else:
        user_form = UserEditForm(instance=user)
        password_form = UserPasswordEditForm(user)

    return render(request, 'logins/edit_user.html', {
        'user_form': user_form,
        'password_form': password_form,
        'user': user
    })

@login_required(login_url='login') 
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dashboard(request):
    if request.user.is_superuser:
        return redirect('admin_dashboard')  

    context = {
        'branch': request.user.branch,
        'department': request.user.department
    }

    # Admin team users get the incharge dashboard
    if request.user.department == 'Admin team':
        template_name = f"logins/incharge_dashboard_{request.user.branch.lower()}.html"
    else:
        template_name = f"logins/employee_dashboard_{request.user.branch.lower()}.html"

    return render(request, template_name, context)



def user_logout(request):
    logout(request)
    return redirect('login')



@login_required(login_url='login')
def purchase_order_report(request):
    orders = PurchaseOrder.objects.all()
    current_year = datetime.now().year
    years = list(range(current_year, current_year - 10, -1))
    return render(request, 'logins/purchase_order_report.html', {'orders': orders,'years': years})



@login_required(login_url='login')
def add_purchase_order(request):
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('purchase_order_report')
    else:
        form = PurchaseOrderForm()

    return render(request, 'logins/add_purchase_order.html', {'form': form})


@login_required(login_url='login')
def edit_purchase_order(request, record_id):
    record = get_object_or_404(PurchaseOrder, id=record_id)
    form = PurchaseOrderForm(request.POST or None, instance=record)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('purchase_order_report')

    return render(request, 'logins/edit_purchase_order.html', {'form': form})



@login_required(login_url='login')
def delete_purchase_order(request, record_id):
    order = get_object_or_404(PurchaseOrder, pk=record_id)
    order.delete()
    messages.success(request, "Purchase order deleted successfully.")
    return redirect('purchase_order_report')


# views.py

@login_required(login_url='login')
def download_purchase_order_report(request):
    month = request.GET.get('month')
    year = request.GET.get('year')

    orders = PurchaseOrder.objects.all()

    if month and year:
        orders = orders.filter(date__year=year, date__month=month)

    if not orders.exists():
        return HttpResponse("No records found.", status=404)

    data = []
    for i, order in enumerate(orders, start=1):
        data.append({
            'S.No': i,
            'DATE': order.date.strftime('%Y-%m-%d'),
            'PO No.': order.po_no,
            'BRANCH': order.branch,
            'DEALER': order.dealer,
            'MATERIAL RECEIVED ON': order.material_received_on.strftime('%Y-%m-%d') if order.material_received_on else '',
            'TOTAL AMOUNT': order.total_amount,
            'PAYMENT DATE': order.payment_date.strftime('%Y-%m-%d') if order.payment_date else '',
            'AMOUNT PAID': order.amount_paid,
            'BALANCE': order.balance_amount,
            'STATUS': order.status,
        })

    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Purchase_Order_Report_{month or "All"}_{year or "All"}.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Report', index=False)

        worksheet = writer.sheets['Report']
        date_style = NamedStyle(name="date_style", number_format='YYYY-MM-DD')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.style = date_style

    return response

# views.py
def dealer_list(request):
    dealers = Dealer.objects.all()
    return render(request, 'logins/dealer_list.html', {'dealers': dealers})


def add_dealer(request):
    form = DealerForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('dealer_list')
    return render(request, 'logins/add_edit_dealer.html', {'form': form, 'action': 'Add'})


def edit_dealer(request, dealer_id):
    dealer = get_object_or_404(Dealer, pk=dealer_id)
    form = DealerForm(request.POST or None, instance=dealer)
    if form.is_valid():
        form.save()
        return redirect('dealer_list')
    return render(request, 'logins/add_edit_dealer.html', {'form': form, 'action': 'Edit'})


def delete_dealer(request, dealer_id):
    dealer = get_object_or_404(Dealer, pk=dealer_id)
    dealer.delete()
    return redirect('dealer_list')


def dealer_purchase_orders(request, dealer_id):
    dealer = get_object_or_404(Dealer, pk=dealer_id)
    orders = DealerPurchaseOrder.objects.filter(dealer=dealer)
    current_year = datetime.now().year
    years = list(range(current_year, current_year - 10, -1))
    return render(request, 'logins/dealer_orders.html', {'dealer': dealer, 'orders': orders,'years': years})


def add_dealer_purchase_order(request, dealer_id):
    dealer = get_object_or_404(Dealer, pk=dealer_id)
    form = DealerPurchaseOrderForm(request.POST or None)
    if form.is_valid():
        order = form.save(commit=False)
        order.dealer = dealer
        order.save()
        return redirect('dealer_purchase_orders', dealer_id=dealer.id)
    return render(request, 'logins/add_edit_order.html', {'form': form, 'action': 'Add', 'dealer': dealer})


def edit_dealer_purchase_order(request, order_id):
    order = get_object_or_404(DealerPurchaseOrder, pk=order_id)
    form = DealerPurchaseOrderForm(request.POST or None, instance=order)
    if form.is_valid():
        form.save()
        return redirect('dealer_purchase_orders', dealer_id=order.dealer.id)
    return render(request, 'logins/add_edit_order.html', {'form': form, 'action': 'Edit', 'dealer': order.dealer})

@login_required(login_url='login')
def edit_dealer_payment(request, payment_id):
    payment = get_object_or_404(DealerPayment, id=payment_id)
    form = DealerPaymentForm(request.POST or None, instance=payment)
    if form.is_valid():
        form.save()
        return redirect('edit_dealer_purchase_order', order_id=payment.order.id)

    return render(request, 'logins/edit_payment.html', {
        'form': form,
        'payment': payment
    })

@login_required(login_url='login')
def delete_dealer_payment(request, payment_id):
    payment = get_object_or_404(DealerPayment, id=payment_id)
    order_id = payment.order.id
    payment.delete()
    return redirect('edit_dealer_purchase_order', order_id=order_id)


def delete_dealer_purchase_order(request, order_id):
    order = get_object_or_404(DealerPurchaseOrder, pk=order_id)
    dealer_id = order.dealer.id
    order.delete()
    return redirect('logins/dealer_purchase_orders', dealer_id=dealer_id)


def add_payment(request, order_id):
    order = get_object_or_404(DealerPurchaseOrder, pk=order_id)
    form = DealerPaymentForm(request.POST or None)
    if form.is_valid():
        payment = form.save(commit=False)
        payment.order = order
        payment.save()
        return redirect('dealer_purchase_orders', dealer_id=order.dealer.id)
    return render(request, 'logins/add_payment.html', {'form': form, 'order': order})


@login_required(login_url='login')
def download_dealer_purchase_order_report(request, dealer_id):
    month = request.GET.get('month')
    year = request.GET.get('year')

    orders = DealerPurchaseOrder.objects.filter(dealer_id=dealer_id)

    if month and year:
        orders = orders.filter(date__year=year, date__month=month)

    if not orders.exists():
        return HttpResponse("No records found.", status=404)

    data = []

    for i, order in enumerate(orders, start=1):
        # Create a formatted string of payment details
        payment_details = []
        for p in order.payments.all():
            payment_details.append(f"Date: {p.date}, Amount: {p.amount}, Balance: {p.balance}")
        payment_details_str = "\n".join(payment_details) if payment_details else "No payments"

        data.append({
            'S.No': i,
            'DATE': order.date.strftime('%Y-%m-%d'),
            'PO No': order.po_no,
            'ABT BRANCH': order.abt_branch,
            'MATERIAL RECEIVED ON': order.material_received_on.strftime('%Y-%m-%d') if order.material_received_on else '',
            'TOTAL AMOUNT': order.total_amount,
            'PAYMENT DETAILS': payment_details_str,
            'PAYMENT STATUS': order.payment_status,
            'GST BILL STATUS': order.gst_bill_status,
        })

    df = pd.DataFrame(data)

    # Create HTTP Response with Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Dealer_PO_Report_{month or 'All'}_{year or 'All'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Report', index=False)

        worksheet = writer.sheets['Report']
        # Optional: Apply styling to make line breaks visible
        for col in worksheet.columns:
            for cell in col:
                cell.alignment = cell.alignment.copy(wrap_text=True)

    return response


@login_required(login_url='login') 
def phd_registration_report(request, branch):
    model_map = {
        "Nagercoil": NagercoilPhdRegistration,
        "Tirunelveli": TirunelveliPhdRegistration,
        "Pudukottai": PudukottaiPhdRegistration,
        "Chennai": ChennaiPhdRegistration
    }

    if branch not in model_map:
        return redirect('dashboard')

    phd_records = model_map[branch].objects.all()
     # Generate available years
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]

    return render(request, "logins/phd_registration_report.html", {
        "branch": branch,
        "phd_records": phd_records,
        'years': years,
    })

model_map = {
    "Nagercoil": NagercoilPhdRegistration,
    "Tirunelveli": TirunelveliPhdRegistration,
    "Pudukottai": PudukottaiPhdRegistration,
    "Chennai": ChennaiPhdRegistration,
}

form_map = {
    "Nagercoil": NagercoilPhdRegistrationForm,
    "Tirunelveli": TirunelveliPhdRegistrationForm,
    "Pudukottai": PudukottaiPhdRegistrationForm,
    "Chennai": ChennaiPhdRegistrationForm,
}

# Add View
@login_required(login_url='login') 
def add_phd_registration(request, branch):
    form_class = form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('phd_registration_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_phd_registration.html", {
        "form": form,
        "branch_name": branch
    })

# Edit View
@login_required(login_url='login') 
def edit_phd_registration(request, branch, record_id):
    if branch not in model_map:
        return redirect('dashboard')

    model = model_map[branch]
    form_class = form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('phd_registration_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_phd_registration.html", {
        "form": form,
        "branch": branch
    })

# Delete View
@login_required(login_url='login') 
def delete_phd_registration(request, branch, record_id):
    if branch not in model_map:
        return redirect('dashboard')

    model = model_map[branch]
    record = get_object_or_404(model, id=record_id)

    # Delete immediately on GET request
    record.delete()
    messages.success(request, "Record deleted successfully.")
    return redirect('phd_registration_report', branch=branch)

project_model_map = {
    "Nagercoil": NagercoilProjectRegistration,
    "Tirunelveli": TirunelveliProjectRegistration,
    "Pudukottai": PudukottaiProjectRegistration,
    "Chennai": ChennaiProjectRegistration,
}

project_form_map = {
    "Nagercoil": NagercoilProjectRegistrationForm,
    "Tirunelveli": TirunelveliProjectRegistrationForm,
    "Pudukottai": PudukottaiProjectRegistrationForm,
    "Chennai": ChennaiProjectRegistrationForm,
}

# Project Registration Report View
@login_required(login_url='login')
def project_registration_report(request, branch):
    if branch not in project_model_map:
        return redirect('dashboard')

    project_records = project_model_map[branch].objects.all()
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]
    return render(request, "logins/project_registration_report.html", {
        "branch": branch,
        "project_records": project_records,
        'years': years,
    })

# Add View
@login_required(login_url='login')
def add_project_registration(request, branch):
    form_class = project_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('project_registration_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_project_registration.html", {
        "form": form,
        "branch_name": branch
    })

# Edit View
@login_required(login_url='login')
def edit_project_registration(request, branch, record_id):
    if branch not in project_model_map:
        return redirect('dashboard')

    model = project_model_map[branch]
    form_class = project_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('project_registration_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_project_registration.html", {
        "form": form,
        "branch": branch
    })

# Delete View
@login_required(login_url='login')
def delete_project_registration(request, branch, record_id):
    if branch not in project_model_map:
        return redirect('dashboard')

    model = project_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    # Immediately delete on GET request
    record.delete()
    messages.success(request, "Project registration record deleted successfully.")
    return redirect('project_registration_report', branch=branch)

intern_model_map = {
    "Nagercoil": NagercoilInternshipRegistration,
    "Tirunelveli": TirunelveliInternshipRegistration,
    "Pudukottai": PudukottaiInternshipRegistration,
    "Chennai": ChennaiInternshipRegistration,
}

intern_form_map = {
    "Nagercoil": NagercoilInternshipRegistrationForm,
    "Tirunelveli": TirunelveliInternshipRegistrationForm,
    "Pudukottai": PudukottaiInternshipRegistrationForm,
    "Chennai": ChennaiInternshipRegistrationForm,
}

@login_required(login_url='login') 
def internship_registration_report(request, branch):
    if branch not in intern_model_map:
        return redirect('dashboard')

    records = intern_model_map[branch].objects.all()
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]
    return render(request, "logins/internship_registration_report.html", {
        "branch": branch,
        "records": records,
        'years': years,
    })

@login_required(login_url='login') 
def add_internship_registration(request, branch):
    form_class = intern_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('internship_registration_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_internship_registration.html", {
        "form": form,
        "branch_name": branch
    })

@login_required(login_url='login') 
def edit_internship_registration(request, branch, record_id):
    if branch not in intern_model_map:
        return redirect('dashboard')

    model = intern_model_map[branch]
    form_class = intern_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('internship_registration_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_internship_registration.html", {
        "form": form,
        "branch": branch
    })

@login_required(login_url='login')
def delete_internship_registration(request, branch, record_id):
    if branch not in intern_model_map:
        return redirect('dashboard')

    model = intern_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    # Delete on GET for script-type delete
    record.delete()
    messages.success(request, "Internship registration record deleted successfully.")
    return redirect('internship_registration_report', branch=branch)

pub_model_map = {
    "Nagercoil": NagercoilPublicationRegistration,
    "Tirunelveli": TirunelveliPublicationRegistration,
    "Pudukottai": PudukottaiPublicationRegistration,
    "Chennai": ChennaiPublicationRegistration,
}

pub_form_map = {
    "Nagercoil": NagercoilPublicationRegistrationForm,
    "Tirunelveli": TirunelveliPublicationRegistrationForm,
    "Pudukottai": PudukottaiPublicationRegistrationForm,
    "Chennai": ChennaiPublicationRegistrationForm,
}

@login_required(login_url='login')
def publication_registration_report(request, branch):
    if branch not in pub_model_map:
        return redirect('dashboard')

    records = pub_model_map[branch].objects.all()
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]
    return render(request, "logins/publication_registration_report.html", {
        "branch": branch,
        "records": records,
        'years': years,
    })

@login_required(login_url='login')
def add_publication_registration(request, branch):
    form_class = pub_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('publication_registration_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_publication_registration.html", {
        "form": form,
        "branch_name": branch
    })

@login_required(login_url='login')
def edit_publication_registration(request, branch, record_id):
    if branch not in pub_model_map:
        return redirect('dashboard')

    model = pub_model_map[branch]
    form_class = pub_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('publication_registration_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_publication_registration.html", {
        "form": form,
        "branch": branch
    })

@login_required(login_url='login')
def delete_publication_registration(request, branch, record_id):
    if branch not in pub_model_map:
        return redirect('dashboard')

    model = pub_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    # Directly delete on GET
    record.delete()
    messages.success(request, "Publication registration record deleted successfully.")
    return redirect('publication_registration_report', branch=branch)




def download_report(request, model, branch, filename):
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Validate month & year
    if not month or not year:
        return HttpResponse("Month and Year parameters are required.", status=400)

    try:
        records = model.objects.filter(date__year=year, date__month=month)
    except ObjectDoesNotExist:
        return HttpResponse("Invalid data request.", status=404)

    if not records.exists():
        return HttpResponse("No records found for the selected month and year.", status=404)

    # Extract fields dynamically from the model
    data = []
    for index, record in enumerate(records, start=1):  # Start numbering from 1
        row = {
            'S.No': index,  # Use sequential numbering
            'Date': record.date.strftime('%Y-%m-%d'),  # Ensure correct date format
            'Reg Code': record.reg_code,
            'Name': record.name,
            'Department': record.department,
            'Type': getattr(record, 'phd_type', '') or getattr(record, 'project_type', '') or 
                    getattr(record, 'internship_type', '') or getattr(record, 'publication_type', ''),
            'College/University or Institution': getattr(record, 'college_university', '') or getattr(record, 'institution', ''),
            'Mobile No': record.mobile_no,
            'Email Id': record.email_id,
            'Amount Paid': record.amount_paid,
            'Amount Balance': record.amount_balance
        }
        data.append(row)

    df = pd.DataFrame(data)

    # Generate Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{branch}_{month}_{year}.xlsx"'

    # Use openpyxl for writing Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Report', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Report']

        # Define a date format style
        date_style = NamedStyle(name="date_style")
        date_style.number_format = 'YYYY-MM-DD'

        # Apply the date format to the Date column
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):  # Date column is now column 2
            for cell in row:
                cell.style = date_style  

    return response


# Define views for different reports
def download_phd_report(request, branch):
    model_map = {
        'Nagercoil': NagercoilPhdRegistration,
        'Tirunelveli': TirunelveliPhdRegistration,
        'Pudukottai': PudukottaiPhdRegistration,
        'Chennai': ChennaiPhdRegistration
    }
    model = model_map.get(branch)
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    return download_report(request, model, branch, 'PhD_Registration_Report')

def download_project_report(request, branch):
    model_map = {
        'Nagercoil': NagercoilProjectRegistration,
        'Tirunelveli': TirunelveliProjectRegistration,
        'Pudukottai': PudukottaiProjectRegistration,
        'Chennai': ChennaiProjectRegistration
    }
    model = model_map.get(branch)
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    return download_report(request, model, branch, 'Project_Registration_Report')

def download_internship_report(request, branch):
    model_map = {
        'Nagercoil': NagercoilInternshipRegistration,
        'Tirunelveli': TirunelveliInternshipRegistration,
        'Pudukottai': PudukottaiInternshipRegistration,
        'Chennai': ChennaiInternshipRegistration
    }
    model = model_map.get(branch)
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    return download_report(request, model, branch, 'Internship_Registration_Report')

def download_publication_report(request, branch):
    model_map = {
        'Nagercoil': NagercoilPublicationRegistration,
        'Tirunelveli': TirunelveliPublicationRegistration,
        'Pudukottai': PudukottaiPublicationRegistration,
        'Chennai': ChennaiPublicationRegistration
    }
    model = model_map.get(branch)
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    return download_report(request, model, branch, 'Publication_Registration_Report')

def download_income_expenditure_report(request, branch):
    model_map = {
        'Nagercoil': NagercoilDailyIncomeExpenditure,
        'Tirunelveli': TirunelveliDailyIncomeExpenditure,
        'Pudukottai': PudukottaiDailyIncomeExpenditure,
        'Chennai': ChennaiDailyIncomeExpenditure
    }
    model = model_map.get(branch)
    
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    
    return download_report_income(request, model, branch, 'Daily_Income_Expenditure_Report')

def download_report_income(request, model, branch, filename):
    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        return HttpResponse("Month and Year parameters are required.", status=400)

    try:
        records = model.objects.filter(date__year=year, date__month=month)
    except ObjectDoesNotExist:
        return HttpResponse("Invalid data request.", status=404)

    if not records.exists():
        return HttpResponse("No records found for the selected month and year.", status=404)

    
    data = []
    for index, record in enumerate(records, start=1): 
        row = {
            'S.No': index,
            'Date': record.date.strftime('%Y-%m-%d'),
            'Description': record.description,
            'Income': record.income,
            'Expense': record.expense,
            'PVC No': record.pvc_no,
            'Balance': record.balance
        }
        data.append(row)

    df = pd.DataFrame(data)

    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{branch}_{month}_{year}.xlsx"'

   
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Report', index=False)

       
        workbook = writer.book
        worksheet = writer.sheets['Report']

        
        date_style = NamedStyle(name="date_style")
        date_style.number_format = 'YYYY-MM-DD'

       
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.style = date_style  

    return response


daily_model_map = {
    "Nagercoil": NagercoilDailyIncomeExpenditure,
    "Tirunelveli": TirunelveliDailyIncomeExpenditure,
    "Pudukottai": PudukottaiDailyIncomeExpenditure,
    "Chennai": ChennaiDailyIncomeExpenditure,
}

# Form mapping for branches
daily_form_map = {
    "Nagercoil": NagercoilDailyIncomeExpenditureForm,
    "Tirunelveli": TirunelveliDailyIncomeExpenditureForm,
    "Pudukottai": PudukottaiDailyIncomeExpenditureForm,
    "Chennai": ChennaiDailyIncomeExpenditureForm,
}

# View Report
@login_required(login_url='login') 
def daily_income_expenditure_report(request, branch):
    if branch not in daily_model_map:
        return redirect('dashboard')

    records = daily_model_map[branch].objects.all()
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]

    return render(request, "logins/daily_income_expenditure_report.html", {
        "branch": branch,
        "records": records,
        "years": years,
    })

# Add Record
@login_required(login_url='login') 
def add_daily_income_expenditure(request, branch):
    form_class = daily_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('daily_income_expenditure_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_daily_income_expenditure.html", {
        "form": form,
        "branch_name": branch
    })

# Edit Record
@login_required(login_url='login') 
def edit_daily_income_expenditure(request, branch, record_id):
    if branch not in daily_model_map:
        return redirect('dashboard')

    model = daily_model_map[branch]
    form_class = daily_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('daily_income_expenditure_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_daily_income_expenditure.html", {
        "form": form,
        "branch": branch
    })

# Delete Record
@login_required(login_url='login') 
def delete_daily_income_expenditure(request, branch, record_id):
    if branch not in daily_model_map:
        return redirect('dashboard')

    model = daily_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    # Allow delete on GET request
    record.delete()
    messages.success(request, "Record deleted successfully.")
    return redirect('daily_income_expenditure_report', branch=branch)

payment_model_map = {
    "Nagercoil": NagercoilPaymentVoucher,
    "Tirunelveli": TirunelveliPaymentVoucher,
    "Pudukottai": PudukottaiPaymentVoucher,
    "Chennai": ChennaiPaymentVoucher,
}

payment_form_map = {
    "Nagercoil": NagercoilPaymentVoucherForm,
    "Tirunelveli": TirunelveliPaymentVoucherForm,
    "Pudukottai": PudukottaiPaymentVoucherForm,
    "Chennai": ChennaiPaymentVoucherForm,
}

# View Payment Voucher Report
@login_required(login_url='login')
def payment_voucher_report(request, branch):
    if branch not in payment_model_map:
        return redirect('dashboard')

    records = payment_model_map[branch].objects.all()
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]

    return render(request, "logins/payment_voucher_report.html", {
        "branch": branch,
        "records": records,
        "years": years,
    })

# Add Payment Voucher
@login_required(login_url='login')
def add_payment_voucher(request, branch):
    form_class = payment_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('payment_voucher_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_payment_voucher.html", {
        "form": form,
        "branch_name": branch
    })

# Edit Payment Voucher
@login_required(login_url='login')
def edit_payment_voucher(request, branch, record_id):
    if branch not in payment_model_map:
        return redirect('dashboard')

    model = payment_model_map[branch]
    form_class = payment_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('payment_voucher_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_payment_voucher.html", {
        "form": form,
        "branch": branch
    })

# Delete Payment Voucher
@login_required(login_url='login')
def delete_payment_voucher(request, branch, record_id):
    if branch not in payment_model_map:
        return redirect('dashboard')

    model = payment_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    # Allow delete directly via GET
    record.delete()
    messages.success(request, "Payment voucher deleted successfully.")
    return redirect('payment_voucher_report', branch=branch)


def download_payment_voucher_report(request, branch):
    model_map = {
        'Nagercoil': NagercoilPaymentVoucher,
        'Tirunelveli': TirunelveliPaymentVoucher,
        'Pudukottai': PudukottaiPaymentVoucher,
        'Chennai': ChennaiPaymentVoucher
    }
    model = model_map.get(branch)
    
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    
    return download_report_payment_voucher(request, model, branch, 'Payment_Voucher_Report')

def download_report_payment_voucher(request, model, branch, filename):
    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        return HttpResponse("Month and Year parameters are required.", status=400)

    try:
        records = model.objects.filter(date__year=year, date__month=month)
    except ObjectDoesNotExist:
        return HttpResponse("Invalid data request.", status=404)

    if not records.exists():
        return HttpResponse("No records found for the selected month and year.", status=404)

    data = []
    for index, record in enumerate(records, start=1): 
        row = {
            'S.No': index,
            'Date': record.date.strftime('%Y-%m-%d'),
            'VC No': record.vc_no,
            'Purpose': record.purpose,
            'Online (A/C)': record.online_payment,  # ✅ Updated field name
            'Cash': record.cash_payment  # ✅ Updated field name
        }
        data.append(row)

    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{branch}_{month}_{year}.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Report', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Report']

        date_style = NamedStyle(name="date_style")
        date_style.number_format = 'YYYY-MM-DD'

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.style = date_style  

    return response



# List of bill types
BILL_TYPES = ["Phd", "Internship", "Project", "Journal", "Sharing", "Patent"]

# Map to get model/form
def get_model(branch, bill_type):
    model_name = f"{branch}{bill_type}Bill"
    return globals()[model_name]

def get_form(branch, bill_type):
    form_name = f"{branch}{bill_type}BillForm"
    return globals()[form_name]

def billwise_dashboard(request, branch):
    return render(request, 'billwise/dashboard.html', {'categories': BILL_TYPES, 'branch': branch})

def bill_list(request, branch, bill_type):
    model = get_model(branch, bill_type)
    records = model.objects.all()
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]
    return render(request, 'billwise/report.html', {'records': records, 'branch': branch, 'bill_type': bill_type,'years': years,})

def add_bill(request, branch, bill_type):
    form_class = get_form(branch, bill_type)
    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('bill_list', branch=branch, bill_type=bill_type)
    else:
        form = form_class()
    return render(request, 'billwise/add_edit.html', {'form': form, 'branch': branch, 'bill_type': bill_type, 'action': 'Add'})

def edit_bill(request, branch, bill_type, record_id):
    model = get_model(branch, bill_type)
    record = get_object_or_404(model, id=record_id)
    form_class = get_form(branch, bill_type)
    if request.method == 'POST':
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('bill_list', branch=branch, bill_type=bill_type)
    else:
        form = form_class(instance=record)
    return render(request, 'billwise/add_edit.html', {'form': form, 'branch': branch, 'bill_type': bill_type, 'action': 'Edit'})

def delete_bill(request, branch, bill_type, record_id):
    model = get_model(branch, bill_type)
    record = get_object_or_404(model, id=record_id)
    record.delete()
    return redirect('bill_list', branch=branch, bill_type=bill_type)



def export_bills_to_excel(request, branch, bill_type):
    model = get_model(branch, bill_type)

    # Get month and year from GET parameters
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Filter by month and year if provided
    queryset = model.objects.all()
    if month and year:
        queryset = queryset.filter(date__month=month, date__year=year)

    response = HttpResponse(content_type='application/ms-excel')
    filename = f"{bill_type}_{branch}_Bills_{month}_{year}.xls" if month and year else f"{bill_type}_{branch}_Bills.xls"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Bills')

    # Header
    columns = ['S.No', 'Date', 'Bill Number', 'Reg No','Name','AMT_RECEIVED', 'mode of payment']
    for col_num, col_name in enumerate(columns):
        ws.write(0, col_num, col_name)

    # Rows
    for row_num, record in enumerate(queryset, start=1):
        ws.write(row_num, 0, row_num)  # S.No
        ws.write(row_num, 1, record.date.strftime("%Y-%m-%d"))
        ws.write(row_num, 2, record.bill_number)
        ws.write(row_num, 3, record.registration_number)
        ws.write(row_num, 4, record.name)
        # ws.write(row_num, 5, float(record.total_amount))
        ws.write(row_num, 5, float(record.cash_received))
        # ws.write(row_num, 7, float(record.online_received))
        ws.write(row_num, 6, record.modeofpayment)
        # ws.write(row_num, 9, float(record.balance))
        # ws.write(row_num, 10, record.payment_status)

    wb.save(response)
    return response

@login_required
def add_work_status(request):
    if request.method == 'POST':
        form = WorkStatusForm(request.POST)
        if form.is_valid():
            status = form.save(commit=False)
            status.user = request.user
            status.branch = request.user.branch
            status.department = request.user.department
            status.save()
            messages.success(request, 'Work status added successfully.')
            return redirect('work_status_list')
    else:
        form = WorkStatusForm()
    return render(request, 'workstatus/add.html', {'form': form})

@login_required
def work_status_list(request):
    statuses = WorkStatus.objects.filter(user=request.user).order_by('-date')
    return render(request, 'workstatus/list.html', {'statuses': statuses})

@login_required
def update_work_status(request, pk):
    status = get_object_or_404(WorkStatus, pk=pk, user=request.user)
    if request.method == 'POST':
        form = WorkStatusForm(request.POST, instance=status)
        if form.is_valid():
            form.save()
            messages.success(request, 'Work status updated.')
            return redirect('work_status_list')
    else:
        form = WorkStatusForm(instance=status)
    return render(request, 'workstatus/update.html', {'form': form})

@login_required
def delete_work_status(request, pk):
    status = get_object_or_404(WorkStatus, pk=pk, user=request.user)
    status.delete()
    messages.success(request, 'Work status deleted.')
    return redirect('work_status_list')
